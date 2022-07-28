import openpyxl
i

class XlUtils:

    def get_rowCnt(self, path, sheet):
        wb = openpyxl.load_workbook(path)
        sheet_obj = wb[sheet]
        return sheet_obj.max_row

    def get_colCnt(self, path, sheet):
        wb = openpyxl.load_workbook(path)
        sheet_obj = wb[sheet]
        return sheet_obj.max_row

    def read_rows(self, path, sheet, row, column):
        wb = openpyxl.load_workbook(path)
        sheet_obj = wb[sheet]
        c = sheet_obj.cell(row=row, column=column)
        return c.value

    def write_to_column(self, path, sheet, row, column, data):
        wb = openpyxl.load_workbook(path)
        sheet_obj = wb[sheet]
        c = sheet_obj.cell(row=row, column=column)
        c.value = data
        wb.save(path)
